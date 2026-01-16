import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import os
from .data_types import CellData, ShapeData, AnchorPoint

class ExcelLoader:
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.cells = []
        self.shapes = []
        
    def load(self):
        self._load_cells()
        self._load_shapes()
        return self.cells, self.shapes

    def _load_cells(self):
        wb = openpyxl.load_workbook(self.filepath, data_only=False) # Keep formulas? Or data_only=True? 
        # Let's keep formulas for now, or maybe value. Diffing formulas might be better.
        # Actually user wants to check difference, usually values matter. 
        # But if formula changes but value is same? 
        # let's stick to default (formulas as strings if possible, or values). 
        # openpyxl default is formulas.
        
        # openpyxl default is formulas.
        
        if self.sheet_name:
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                raise ValueError(f"Sheet '{self.sheet_name}' not found in {self.filepath}")
        else:
            ws = wb.active # Assume first sheet for now
        
        for row in ws.iter_rows():
            for cell in row:
                self.cells.append(CellData(
                    row=cell.row,
                    col=cell.column,
                    value=cell.value,
                    coordinate=cell.coordinate
                ))
        wb.close()

    def _load_shapes(self):
        # We need to map worksheet relationships to find the drawing file
        # For simplicity in V1, we iterate all drawing XMLs found in the zip 
        # and assume they belong to the active sheet or catch them all.
        # A more robust way involves parsing xl/worksheets/sheet1.xml to find the drawing rId.
        
        ns = {
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
        }
        
        with zipfile.ZipFile(self.filepath, 'r') as z:
            # TODO: Filter shapes by sheet_name if possible. 
            # Currently parses all drawings in the file.
            # This might include shapes from other sheets.
            # Improvement: parsing worksheet relationships to find specific drawing file.
            
            # Simple heuristic: look for drawing files
            drawing_files = [f for f in z.namelist() if 'xl/drawings/drawing' in f]
            
            for df in drawing_files:
                with z.open(df) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    
                    # twoCellAnchor is the most common for shapes placed in grid
                    for anchor in root.findall('.//xdr:twoCellAnchor', ns):
                        self._parse_anchor_shape(anchor, ns)
                        
                    # oneCellAnchor (less common for main shapes, often for comments/buttons)
                    for anchor in root.findall('.//xdr:oneCellAnchor', ns):
                        self._parse_anchor_shape(anchor, ns)

    def _parse_anchor_shape(self, anchor, ns):
        # Get Shape Info
        sp = anchor.find('xdr:sp', ns)
        if sp is None:
            # specific case for textboxes might be under 'xdr:sp' normally,
            # but groups or other types exist. 
            # If no 'sp', check for graphicFrame (charts) or grpSp (groups)
            # For this MVP, we focus on 'sp' (Shape)
            return 

        nvSpPr = sp.find('xdr:nvSpPr', ns)
        cNvPr = nvSpPr.find('xdr:cNvPr', ns)
        shape_id = cNvPr.get('id')
        shape_name = cNvPr.get('name')
        
        # Get From Anchor
        fr = anchor.find('xdr:from', ns)
        from_pt = self._extract_point(fr, ns)
        
        # Get To Anchor (only for twoCellAnchor)
        to = anchor.find('xdr:to', ns)
        to_pt = self._extract_point(to, ns) if to is not None else None
        
        # Get Text content (if any)
        text_content = ""
        txBody = sp.find('xdr:txBody', ns)
        if txBody:
            # Extract all text paragraphs
            paragraphs = txBody.findall('.//a:p//a:t', ns)
            text_content = "\n".join([t.text for t in paragraphs if t.text])

        self.shapes.append(ShapeData(
            id=shape_id,
            name=shape_name,
            type_name="Shape", # Simplified
            from_anchor=from_pt,
            to_anchor=to_pt,
            text=text_content
        ))

    def _extract_point(self, node, ns) -> AnchorPoint:
        col = int(node.find('xdr:col', ns).text)
        colOff = int(node.find('xdr:colOff', ns).text)
        row = int(node.find('xdr:row', ns).text)
        rowOff = int(node.find('xdr:rowOff', ns).text)
        # Note: XML rows/cols are 0-indexed usually in drawings, but Excel UI is 1-indexed.
        # OpenPyxl is 1-indexed.
        # Let's verify this with a test. Commonly drawingML is 0-indexed.
        return AnchorPoint(row=row, col=col, row_off=rowOff, col_off=colOff)
