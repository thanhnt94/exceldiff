import customtkinter as ctk
import win32com.client as win32
import os
import threading
import pythoncom
import difflib
import openpyxl
from openpyxl.utils import get_column_letter 
from dataclasses import dataclass
from typing import List, Dict, Tuple, Any
from tkinter import filedialog, messagebox
from datetime import datetime
import time 

# --- CONFIGURATION & CONSTANTS ---
AUTHOR_ID = "KNT15083"
PROJECT_NAME = "ExcelDiff Tool"
THEME_COLOR = "#2C3E50"  
ACCENT_COLOR = "#27AE60" 
HOVER_COLOR = "#2ECC71"
TOLERANCE_DEFAULT = 3 

# UI Colors & Fonts
STATUS_BAR_BG = "#23272D" 
CARD_BORDER_COLOR = "#3E4C5E"
MAIN_FONT = ("Segoe UI", 12)
BOLD_FONT = ("Segoe UI", 12, "bold")
HEADER_FONT = ("Segoe UI", 13, "bold")

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# Excel Constants
xlUp = -4162
xlSrcRange = 1
xlYes = 1
xlTop = -4160
xlLeft = -4131
xlNormalView = 1 
xlCalculationManual = -4135
xlCalculationAutomatic = -4105
xlMaximized = -4137

# --- LOCALIZATION DATA ---
LANGUAGES = {
    "English": {
        "source": "Source (Old):",
        "target": "Target (New):",
        "browse": "📂 Browse",
        "open": "↗",
        "placeholder_old": "Path to original file...",
        "placeholder_new": "Path to new file...",
        "output": "Output Folder:",
        "placeholder_out": "Default: Downloads", 
        "limit_scan": "Limit Scan Range",
        "scope_types": ["Whole Sheet", "Columns Only", "Rows Only", "Specific Range"],
        "force_kill": "Force Kill Excel",
        "start_btn": "START COMPARISON",
        "processing": "⏳ PROCESSING...",
        "status_ready": " Ready.",
        "status_done": "Done! Ready for next task.",
        "status_error": "Error occurred.",
        "msg_success": "Comparison Complete!\nSaved to: ",
        "msg_error": "An error occurred:\n",
        "ph_range_col": "e.g. A:D",
        "ph_range_row": "e.g. 1:100",
        "ph_range_spec": "e.g. A1:H50",
        "sel_file_first": "Select File First",
        "help_btn": "❓ Help",
        "help_title": "User Guide",
        "help_content": """HOW TO USE EXDIFF TOOL:

1. Select Files:
   - Click 'Browse' to select the Source (Old) and Target (New) Excel files.
   - Wait for the sheet list to load, then select the specific sheet to compare.

2. Configuration:
   - Limit Scan Range: Use this if you only want to compare a specific area.
   - Report Mode: Choose 'Full Report' or 'Changes Only'.
   - Force Kill Excel: Kills all Excel processes before starting.
   - Highlight Changes: Red background for cells containing modified shapes.

3. Run:
   - Click 'START COMPARISON'.""",
        "lbl_report_mode": "Report Mode:",
        "report_modes": ["Full Report", "Changes Only"],
        "chk_highlight": "Highlight Changes"
    },
    "Tiếng Việt": {
        "source": "Nguồn (Cũ):",
        "target": "Đích (Mới):",
        "browse": "📂 Chọn File",
        "open": "↗",
        "placeholder_old": "Đường dẫn file gốc...",
        "placeholder_new": "Đường dẫn file mới...",
        "output": "Thư mục ra:",
        "placeholder_out": "Mặc định: Downloads", 
        "limit_scan": "Giới hạn vùng quét",
        "scope_types": ["Toàn bộ Sheet", "Chỉ Cột", "Chỉ Hàng", "Vùng cụ thể"],
        "force_kill": "Force Kill Excel",
        "start_btn": "BẮT ĐẦU SO SÁNH",
        "processing": "⏳ ĐANG XỬ LÝ...",
        "status_ready": " Sẵn sàng.",
        "status_done": "Hoàn tất! Sẵn sàng.",
        "status_error": "Có lỗi xảy ra.",
        "msg_success": "So sánh hoàn tất!\nĐã lưu tại: ",
        "msg_error": "Đã xảy ra lỗi:\n",
        "ph_range_col": "VD: A:D",
        "ph_range_row": "VD: 1:100",
        "ph_range_spec": "VD: A1:H50",
        "sel_file_first": "Chọn file trước",
        "help_btn": "❓ Hướng dẫn",
        "help_title": "Hướng dẫn sử dụng",
        "help_content": """CÁCH SỬ DỤNG EXDIFF TOOL:

1. Chọn File:
   - Nhấn 'Chọn File' để chọn file Excel cũ (Source) và mới (Target).
   - Chờ danh sách sheet tải xong, sau đó chọn sheet cần so sánh.

2. Cấu hình:
   - Giới hạn vùng quét: Tích chọn nếu chỉ muốn so sánh vùng cụ thể.
   - Chế độ báo cáo: 'Báo cáo đầy đủ' hoặc 'Chỉ thay đổi'.
   - Force Kill Excel: Tắt cưỡng bức mọi tiến trình Excel (Dùng khi bị treo).
   - Tô đỏ thay đổi: Nếu tích, ô chứa shape bị thay đổi sẽ được tô nền đỏ.

3. Chạy:
   - Nhấn 'BẮT ĐẦU SO SÁNH'.""",
        "lbl_report_mode": "Chế độ báo cáo:",
        "report_modes": ["Báo cáo đầy đủ", "Chỉ thay đổi"],
        "chk_highlight": "Tô đỏ thay đổi"
    },
    "日本語": {
        "source": "元ファイル (旧):",
        "target": "対象ファイル (新):",
        "browse": "📂 参照",
        "open": "↗",
        "placeholder_old": "元ファイルのパス...",
        "placeholder_new": "新ファイルのパス...",
        "output": "出力フォルダ:",
        "placeholder_out": "デフォルト: Downloads", 
        "limit_scan": "スキャン範囲を制限",
        "scope_types": ["シート全体", "列のみ", "行のみ", "指定範囲"],
        "force_kill": "Force Kill Excel", 
        "start_btn": "比較開始",
        "processing": "⏳ 処理中...",
        "status_ready": " 準備完了。",
        "status_done": "完了しました。",
        "status_error": "エラーが発生しました。",
        "msg_success": "比較完了！\n保存先: ",
        "msg_error": "エラーが発生しました:\n",
        "ph_range_col": "例: A:D",
        "ph_range_row": "例: 1:100",
        "ph_range_spec": "例: A1:H50",
        "sel_file_first": "ファイルを選択してください",
        "help_btn": "❓ ヘルプ",
        "help_title": "ユーザーガイド",
        "help_content": """EXDIFF ツールの使い方:

1. ファイル選択:
   - 「参照」をクリックして、比較元（旧）と対象（新）を選択します。
   - シート一覧から比較対象を選択してください。

2. 設定:
   - スキャン範囲を制限: 特定の範囲のみを比較したい場合に使用します。
   - レポートモード: 「全レポート」または「変更のみ」。
   - Force Kill Excel: 開始前にExcelプロセスを強制終了します。
   - 変更をハイライト: チェックすると、変更された図形のあるセルが赤く塗りつぶされます。

3. 実行:
   - 「比較開始」をクリックします。""",
        "lbl_report_mode": "レポートモード:",
        "report_modes": ["全レポート", "変更のみ"],
        "chk_highlight": "変更箇所を赤く表示"
    }
}

# --- DATA STRUCTURES ---

@dataclass
class ShapeData:
    id: int
    name: str
    height: float
    width: float
    abs_top: float
    abs_left: float
    anchor_address: str
    anchor_row: int
    anchor_col: int
    rel_top: float  
    rel_left: float 

@dataclass
class CellDiff:
    index: int
    category: str 
    action: str   
    address_id: str
    old_val: Any
    new_val: Any
    details: str

@dataclass
class ShapeDiff:
    index: int
    shape_id: int
    name: str
    verdict: str
    diff_x: float
    diff_y: float
    diff_w: float
    diff_h: float
    old_anchor: str
    exp_anchor: str
    act_anchor: str
    old_rel_x: float
    old_rel_y: float
    new_rel_x: float
    new_rel_y: float

# --- CORE LOGIC ENGINE ---

class ExcelEngine:
    def __init__(self):
        self.app = None
        self.wb_old = None
        self.wb_new = None

    def kill_excel(self):
        """Force kills Excel processes."""
        try:
            os.system("taskkill /f /im excel.exe")
        except Exception:
            pass

    def robust_open(self, path_old, path_new, force_reset=True):
        """
        Opens Excel safely and loads workbooks in a SEPARATE instance.
        Sử dụng DispatchEx để tránh ảnh hưởng đến các file Excel đang mở của người dùng.
        """
        if force_reset:
            self.kill_excel()
        
        # Initialize COM for this thread
        pythoncom.CoInitialize() 
        
        try:
            # [FIXED] Sửa lỗi cú pháp: Bỏ .client bị thừa
            # win32 ở đây chính là win32com.client (do bạn import as win32)
            try:
                self.app = win32.DispatchEx("Excel.Application")
            except Exception:
                # Fallback nếu máy không hỗ trợ DispatchEx (hiếm)
                self.app = win32.Dispatch("Excel.Application")
            
            # Cấu hình ẩn và tắt cảnh báo cho instance riêng biệt này
            self.app.Visible = False
            self.app.DisplayAlerts = False
            self.app.AskToUpdateLinks = False
            
            # Tối ưu: Tắt cập nhật màn hình ngay từ đầu
            self.app.ScreenUpdating = False 
            
            # --- MỞ FILE CŨ (SOURCE/OLD) ---
            try:
                self.wb_old = self.app.Workbooks.Open(
                    Filename=os.path.abspath(path_old), 
                    ReadOnly=True, 
                    UpdateLinks=0, 
                    CorruptLoad=1
                )
                
                # Xử lý Protected View
                if self.app.ProtectedViewWindows.Count > 0:
                    try:
                        for i in range(self.app.ProtectedViewWindows.Count, 0, -1):
                            self.app.ProtectedViewWindows(i).Edit()
                    except: pass

                # Xử lý Sheet bị khóa
                if self.wb_old.ActiveSheet.ProtectContents:
                    try: self.wb_old.ActiveSheet.Unprotect()
                    except: pass
            
            except Exception as e:
                raise Exception(f"Error opening Source file (Old): {str(e)}")
            
            # --- MỞ FILE MỚI (TARGET/NEW) ---
            try:
                self.wb_new = self.app.Workbooks.Open(
                    Filename=os.path.abspath(path_new), 
                    ReadOnly=True, 
                    UpdateLinks=0, 
                    CorruptLoad=1
                )
                
                if self.app.ProtectedViewWindows.Count > 0:
                    try:
                        for i in range(self.app.ProtectedViewWindows.Count, 0, -1):
                            self.app.ProtectedViewWindows(i).Edit()
                    except: pass

                if self.wb_new.ActiveSheet.ProtectContents:
                    try: self.wb_new.ActiveSheet.Unprotect()
                    except: pass

            except Exception as e:
                raise Exception(f"Error opening Target file (New): {str(e)}")
                
        except Exception as e:
            # Nếu gặp lỗi, cleanup để tắt process Excel ẩn
            self.cleanup()
            raise Exception(f"System Error: {str(e)}")

    def get_sheet_by_name(self, wb, name):
        try:
            return wb.Sheets(name)
        except:
            return wb.Sheets(1) 

    def extract_shapes(self, ws, file_label="Unknown File", scan_range_addr=None, log_func=None) -> Dict[int, ShapeData]:
        # [CRITICAL LOGIC UPDATE] 
        # Sử dụng wb.Windows(1) thay vì ActiveWindow để tránh lỗi focus nhầm file
        
        try:
            if log_func: log_func(f"[{file_label}] Preparing View State...")

            # 0. HANDLE PROTECTED VIEW
            if self.app.ProtectedViewWindows.Count > 0:
                try:
                    count = self.app.ProtectedViewWindows.Count
                    if log_func: log_func(f"[{file_label}] Detect {count} Protected View windows. Attempting to unblock...")
                    for i in range(count, 0, -1):
                        self.app.ProtectedViewWindows(i).Edit()
                except: pass

            # 1. Access Specific Window of the Workbook (Safe Targeting)
            wb = ws.Parent
            
            if wb.Windows.Count == 0:
                wb.NewWindow()
                
            my_window = wb.Windows(1) 
            my_window.Visible = True
            
            wb.Activate()
            ws.Activate()
            
            # [TRICK] Bật cập nhật màn hình tạm thời
            self.app.ScreenUpdating = True 
            
            # 2. Force View & Zoom trên đúng cửa sổ đó
            try:
                if my_window.View != xlNormalView:
                    my_window.View = xlNormalView
            except: pass
            
            # 3. Toggle Zoom with Retry
            for attempt in range(3):
                try:
                    my_window.Zoom = 100
                    if my_window.Zoom == 100: break
                    
                    # Vẩy zoom
                    my_window.Zoom = 101
                    my_window.Zoom = 100
                    if my_window.Zoom == 100: break
                except:
                    time.sleep(0.5)
            
            # 4. VERIFY 
            current_zoom = my_window.Zoom
            if log_func: log_func(f"[{file_label}] View State Verified: Zoom {current_zoom}%")
            
            if abs(current_zoom - 100) > 1: 
                err_msg = (f"[{file_label}] CRITICAL: Failed to enforce 100% Zoom. Current is {current_zoom}%.")
                raise Exception(err_msg)
                
        except Exception as e:
            err_msg = f"[{file_label}] CRITICAL VIEW ERROR: {str(e)}"
            print(err_msg)
            if log_func: log_func(err_msg)
            raise Exception(err_msg)

        # --- [SPEED OPTIMIZATION 1] ---
        # Tắt cập nhật màn hình NGAY LẬP TỨC sau khi đã fix xong Zoom
        self.app.ScreenUpdating = False 

        # 5. SCAN VALUES
        if log_func: log_func(f"[{file_label}] Starting Shape Scan...")

        shape_dict = {}
        shapes = ws.Shapes
        total_shapes = shapes.Count 
        
        # --- [SPEED OPTIMIZATION 2] ---
        r_min, c_min, r_max, c_max = 0, 0, 0, 0
        use_custom_bounds = False
        
        try:
            if scan_range_addr and scan_range_addr.strip() != "":
                scan_area = ws.Range(scan_range_addr)
                r_min = scan_area.Row
                c_min = scan_area.Column
                r_max = r_min + scan_area.Rows.Count - 1
                c_max = c_min + scan_area.Columns.Count - 1
                use_custom_bounds = True
            else:
                scan_area = ws.UsedRange
                r_min = scan_area.Row
                c_min = scan_area.Column
                r_max = r_min + scan_area.Rows.Count - 1
                c_max = c_min + scan_area.Columns.Count - 1
                use_custom_bounds = True 
        except Exception as e:
            if log_func: log_func(f"[{file_label}] Warning: Bound calc failed. Scan all.")
            use_custom_bounds = False
        
        for i, shp in enumerate(shapes):
            current_idx = i + 1
            if log_func and current_idx % 10 == 0: 
                log_func(f"[{file_label}] Scanning Shape {current_idx}/{total_shapes}...")

            try:
                # Check Bounds bằng số học
                tl_cell = shp.TopLeftCell
                if use_custom_bounds:
                    if not (r_min <= tl_cell.Row <= r_max and c_min <= tl_cell.Column <= c_max):
                        continue # Bỏ qua nếu nằm ngoài vùng
                
                s_data = ShapeData(
                    id=shp.ID,
                    name=shp.Name,
                    height=shp.Height,
                    width=shp.Width,
                    abs_top=shp.Top,
                    abs_left=shp.Left,
                    anchor_address=tl_cell.Address,
                    anchor_row=tl_cell.Row,
                    anchor_col=tl_cell.Column,
                    rel_top=shp.Top - tl_cell.Top,   
                    rel_left=shp.Left - tl_cell.Left 
                )
                shape_dict[shp.ID] = s_data
            except Exception:
                continue 
        
        if log_func: log_func(f"[{file_label}] Scan Complete. Found {len(shape_dict)} valid shapes.")
        return shape_dict

    def get_used_range_values(self, ws):
        data = ws.UsedRange.Value
        if data is None: return []
        if not isinstance(data, tuple): return [[data]] 
        return [list(row) for row in data]

    def _optimize_speed(self, enable=True):
        """[UPDATED] Tắt/Bật cập nhật màn hình để tăng tốc độ Copy."""
        try:
            if enable:
                self.app.ScreenUpdating = False
                self.app.Calculation = xlCalculationManual 
                self.app.EnableEvents = False
            else:
                self.app.ScreenUpdating = True
                self.app.Calculation = xlCalculationAutomatic 
                self.app.EnableEvents = True
        except:
            pass

    def create_report_workbook(self, output_folder, cell_diffs: List[CellDiff], shape_diffs: List[ShapeDiff], 
                               ws_src_old, ws_src_new, only_diffs=False, highlight_changes=False):
        """Generates the formatted Excel report."""
        
        self._optimize_speed(True)
        
        try:
            wb_out = self.app.Workbooks.Add()
            
            # 1. COPY SOURCE SHEETS
            try:
                ws_src_old.Parent.Activate()
                ws_src_old.Activate()
            except: pass
            
            ws_src_old.Copy(After=wb_out.Sheets(wb_out.Sheets.Count))
            ws_copy_old = wb_out.Sheets(wb_out.Sheets.Count)
            ws_copy_old.Name = "Source_Old"
            
            try:
                ws_src_new.Parent.Activate()
                ws_src_new.Activate()
            except: pass
            
            ws_src_new.Copy(After=wb_out.Sheets(wb_out.Sheets.Count))
            ws_copy_new = wb_out.Sheets(wb_out.Sheets.Count)
            ws_copy_new.Name = "Source_New"
            
            # --- TÔ MÀU CẢ 2 SHEET ---
            if highlight_changes:
                red_addresses_new = []
                red_addresses_old = []
                
                for s in shape_diffs:
                    if s.verdict != "MATCH" and "DELETED" not in s.verdict:
                        if s.act_anchor and s.act_anchor != "N/A":
                            red_addresses_new.append(s.act_anchor)
                            
                    if s.verdict != "MATCH" and "NEW" not in s.verdict:
                        if s.old_anchor and s.old_anchor != "N/A":
                            red_addresses_old.append(s.old_anchor)
                
                # Tô màu cho New Sheet
                if red_addresses_new:
                    chunk_size = 20
                    for i in range(0, len(red_addresses_new), chunk_size):
                        chunk = red_addresses_new[i:i + chunk_size]
                        addr_str = ",".join(chunk)
                        try:
                            ws_copy_new.Range(addr_str).Interior.Color = 255
                        except: pass
                        
                # Tô màu cho Old Sheet
                if red_addresses_old:
                    chunk_size = 20
                    for i in range(0, len(red_addresses_old), chunk_size):
                        chunk = red_addresses_old[i:i + chunk_size]
                        addr_str = ",".join(chunk)
                        try:
                            ws_copy_old.Range(addr_str).Interior.Color = 255
                        except: pass

            # --- Sheet 1: Cell_Grid_Report ---
            ws1 = wb_out.Sheets(1)
            ws1.Name = "Cell_Grid_Report"
            
            headers1 = ["Index", "Category", "Action", "Address / ID", "Old Value / Size", "New Value / Size", "Details"]
            data1 = [[d.index, d.category, d.action, d.address_id, str(d.old_val), str(d.new_val), d.details] for d in cell_diffs]
                
            if data1:
                ws1.Range(ws1.Cells(1, 1), ws1.Cells(1, 7)).Value = headers1
                rng_data = ws1.Range(ws1.Cells(2, 1), ws1.Cells(len(data1)+1, 7))
                rng_data.Value = data1
                
                last_row = len(data1) + 1
                ws1.ListObjects.Add(xlSrcRange, ws1.Range(f"A1:G{last_row}"), 0, xlYes).TableStyle = "TableStyleMedium2"
                
                try: ws1.Cells.Font.Name = "Meiryo UI"
                except: pass

                # Hyperlinks
                for i, d in enumerate(cell_diffs):
                    row_idx = i + 2
                    cell_addr_obj = ws1.Cells(row_idx, 4) 
                    action = d.action
                    addr = d.address_id
                    
                    target_sheet = "Source_New"
                    if "DELETED" in action:
                        target_sheet = "Source_Old"
                    
                    if ":" not in addr and "Row" not in addr and " " not in addr: 
                        sub_addr = f"'{target_sheet}'!{addr}"
                        try:
                            ws1.Hyperlinks.Add(Anchor=cell_addr_obj, Address="", SubAddress=sub_addr, TextToDisplay=addr)
                        except: pass
                    elif "Row" in addr:
                        try:
                            row_num = addr.replace("Row", "").strip()
                            sub_addr = f"'{target_sheet}'!A{row_num}"
                            ws1.Hyperlinks.Add(Anchor=cell_addr_obj, Address="", SubAddress=sub_addr, TextToDisplay=addr)
                        except: pass
                
                ws1.Columns.AutoFit()
            else:
                ws1.Cells(1, 1).Value = "No Cell/Grid Differences Found."

            # --- Sheet 2: Shape_Report ---
            ws2 = wb_out.Sheets.Add(After=ws1)
            ws2.Name = "Shape_Report"
            
            headers2 = ["Index", "Shape ID", "Shape Name", "Verdict", 
                        "Diff X", "Diff Y", "Diff W", "Diff H", 
                        "Old Anchor", "New Anchor (Exp)", "New Anchor (Act)",
                        "Old Rel X", "Old Rel Y", "New Rel X", "New Rel Y"]
            
            # FILTER DATA BASED ON REPORT MODE
            final_shape_diffs = shape_diffs
            if only_diffs:
                final_shape_diffs = [s for s in shape_diffs if s.verdict != "MATCH"]

            data2 = []
            for s in final_shape_diffs:
                data2.append([
                    s.index, s.shape_id, s.name, s.verdict, 
                    round(s.diff_x, 2), round(s.diff_y, 2), round(s.diff_w, 2), round(s.diff_h, 2),
                    s.old_anchor, s.exp_anchor, s.act_anchor,
                    round(s.old_rel_x, 2), round(s.old_rel_y, 2), 
                    round(s.new_rel_x, 2), round(s.new_rel_y, 2)
                ])
                
            if data2:
                ws2.Range(ws2.Cells(1, 1), ws2.Cells(1, 15)).Value = headers2
                rng_data2 = ws2.Range(ws2.Cells(2, 1), ws2.Cells(len(data2)+1, 15))
                rng_data2.Value = data2
                
                last_row = len(data2) + 1
                ws2.ListObjects.Add(xlSrcRange, ws2.Range(f"A1:O{last_row}"), 0, xlYes).TableStyle = "TableStyleMedium2"
                
                try: ws2.Cells.Font.Name = "Meiryo UI"
                except: pass

                # Hyperlinks
                for i, s in enumerate(final_shape_diffs):
                    row_idx = i + 2
                    
                    if s.old_anchor and s.old_anchor != "N/A":
                        cell_old = ws2.Cells(row_idx, 9)
                        sub_addr_old = f"'Source_Old'!{s.old_anchor.replace('$','')}"
                        try:
                            ws2.Hyperlinks.Add(Anchor=cell_old, Address="", SubAddress=sub_addr_old, TextToDisplay=s.old_anchor)
                        except: pass
                        
                    if s.act_anchor and s.act_anchor != "N/A":
                        cell_new = ws2.Cells(row_idx, 11)
                        sub_addr_new = f"'Source_New'!{s.act_anchor.replace('$','')}"
                        try:
                            ws2.Hyperlinks.Add(Anchor=cell_new, Address="", SubAddress=sub_addr_new, TextToDisplay=s.act_anchor)
                        except: pass

                ws2.Columns.AutoFit()
            else:
                ws2.Cells(1, 1).Value = "No Shapes Found (or all Matched)."

            # =========================================================================
            # [CẬP NHẬT] ĐẶT TÊN FILE THEO ĐỊNH DẠNG MỚI
            # Format: ExcelDiff_[10 kí tự file gốc]_[Tên sheet gốc]_[Thời gian].xlsx
            # =========================================================================
            try:
                # 1. Lấy tên file gốc (Source) và bỏ phần mở rộng (.xlsx)
                src_filename = ws_src_old.Parent.Name
                src_name_clean = os.path.splitext(src_filename)[0]
                
                # 2. Cắt lấy 10 ký tự đầu tiên
                short_src_name = src_name_clean[:10].strip()
                
                # 3. Lấy tên Sheet gốc
                sheet_name = ws_src_old.Name
                # Xử lý ký tự đặc biệt trong tên sheet (để an toàn khi lưu file)
                safe_sheet_name = "".join([c if c.isalnum() or c in " -_" else "_" for c in sheet_name])

                # 4. Lấy thời gian
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # 5. Tạo tên file cuối cùng
                out_name = f"ExcelDiff_{short_src_name}_{safe_sheet_name}_{timestamp}.xlsx"
                
            except Exception:
                # Fallback nếu lỗi lấy tên
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_name = f"ExcelDiff_{timestamp}.xlsx"

            out_path = os.path.join(output_folder, out_name)
            
            if os.path.exists(out_path):
                try: os.remove(out_path)
                except: pass
                
            wb_out.SaveAs(out_path)
            return out_path
        finally:
            self._optimize_speed(False)

    def cleanup(self):
        try:
            if self.wb_old: self.wb_old.Close(False)
            if self.wb_new: self.wb_new.Close(False)
            self.app.Quit()
        except:
            pass

class Comparator:
    def __init__(self, tolerance):
        self.tolerance = tolerance

    def compare_grids_and_cells(self, ws_old, ws_new, engine: ExcelEngine, log_func=None) -> Tuple[List[CellDiff], Dict[int, int], Dict[int, int]]:
        report = []
        
        if log_func: log_func("Reading Grid Data...")
        raw_old = engine.get_used_range_values(ws_old)
        raw_new = engine.get_used_range_values(ws_new)
        
        if log_func: log_func(f"Analyzing Grid Structure (Rows: {len(raw_old)} vs {len(raw_new)})...")
        
        # [UPDATED] TỐI ƯU HIỆU SUẤT: Dùng HASH để so sánh cấu trúc dòng nhanh hơn thay vì string
        sig_old = [hash(tuple(r)) for r in raw_old]
        sig_new = [hash(tuple(r)) for r in raw_new]
        
        # [UPDATED] autojunk=False để tăng độ chính xác với dữ liệu số
        matcher = difflib.SequenceMatcher(None, sig_old, sig_new, autojunk=False)
        
        row_map = {} 
        idx_counter = 1
        
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                for k in range(i2 - i1):
                    row_map[i1 + k + 1] = j1 + k + 1
            elif tag == 'delete':
                for k in range(i2 - i1):
                    report.append(CellDiff(idx_counter, "ROW", "DELETED", f"Row {i1+k+1}", "", "", "Row removed"))
                    idx_counter += 1
            elif tag == 'insert':
                for k in range(j2 - j1):
                    report.append(CellDiff(idx_counter, "ROW", "INSERTED", f"Row {j1+k+1}", "", "", "Row added"))
                    idx_counter += 1
            elif tag == 'replace':
                if (i2-i1) == (j2-j1):
                    for k in range(i2-i1):
                        row_map[i1+k+1] = j1+k+1
        
        col_map = {i: i for i in range(1, 256)} 
        
        # --- [NEW FEATURE] CHECK COLUMN WIDTH ---
        if log_func: log_func("Checking Column Widths...")
        try:
            # Lấy số cột lớn nhất để quét
            max_col_check = 0
            try: max_col_check = max(ws_old.UsedRange.Columns.Count, ws_new.UsedRange.Columns.Count)
            except: pass
            
            if max_col_check > 0:
                for c_idx in range(1, max_col_check + 1):
                    try:
                        w_old = ws_old.Columns(c_idx).ColumnWidth
                        w_new = ws_new.Columns(c_idx).ColumnWidth
                        
                        # So sánh với sai số 0.1 (do float point)
                        if abs(w_old - w_new) > 0.1:
                            col_letter = get_column_letter(c_idx)
                            report.append(CellDiff(
                                idx_counter, "COLUMN", "RESIZED", 
                                f"Col {col_letter}", 
                                round(w_old, 2), round(w_new, 2), 
                                "Width changed"
                            ))
                            idx_counter += 1
                    except: continue
        except Exception as e:
            if log_func: log_func(f"Warning: Column Width check failed ({str(e)})")

        # --- [NEW FEATURE] CHECK ROW HEIGHT ---
        if log_func: log_func("Checking Row Heights...")
        try:
            max_row_check = 0
            try: max_row_check = max(ws_old.UsedRange.Rows.Count, ws_new.UsedRange.Rows.Count)
            except: pass
            
            if max_row_check > 0:
                for r_idx in range(1, max_row_check + 1):
                    try:
                        h_old = ws_old.Rows(r_idx).RowHeight
                        h_new = ws_new.Rows(r_idx).RowHeight
                        
                        if abs(h_old - h_new) > 0.1:
                            report.append(CellDiff(
                                idx_counter, "ROW", "RESIZED", 
                                f"Row {r_idx}", 
                                round(h_old, 2), round(h_new, 2), 
                                "Height changed"
                            ))
                            idx_counter += 1
                    except: continue
        except Exception as e:
            if log_func: log_func(f"Warning: Row Height check failed ({str(e)})")

        total_rows_check = len(raw_old)
        for r_old_idx, row_data in enumerate(raw_old):
            if log_func and r_old_idx % 100 == 0:
                log_func(f"Comparing Cells Row: {r_old_idx+1}/{total_rows_check}...")

            excel_r_old = r_old_idx + 1
            if excel_r_old in row_map:
                excel_r_new = row_map[excel_r_old]
                if excel_r_new - 1 < len(raw_new):
                    row_new_data = raw_new[excel_r_new - 1]
                    max_col = max(len(row_data), len(row_new_data))
                    for c_idx in range(max_col):
                        val_old = row_data[c_idx] if c_idx < len(row_data) else None
                        val_new = row_new_data[c_idx] if c_idx < len(row_new_data) else None
                        
                        # [UPDATED] TĂNG ĐỘ CHÍNH XÁC: So sánh theo Type thay vì String
                        is_diff = False
                        
                        # Chuẩn hóa None về "" để so sánh an toàn
                        v1_check = "" if val_old is None else val_old
                        v2_check = "" if val_new is None else val_new
                        
                        # Logic so sánh thông minh
                        if type(v1_check) != type(v2_check):
                            # Nếu khác kiểu, nhưng là số (VD: 10 vs 10.0) -> check sai số
                            if isinstance(v1_check, (int, float)) and isinstance(v2_check, (int, float)):
                                if abs(float(v1_check) - float(v2_check)) > 0.000001:
                                    is_diff = True
                            else:
                                # Thử ép kiểu string nếu không phải số thuần túy
                                if str(v1_check).strip() != str(v2_check).strip():
                                    is_diff = True
                        else:
                            # Cùng kiểu dữ liệu
                            if isinstance(v1_check, (int, float)):
                                if abs(v1_check - v2_check) > 0.000001: # Epsilon cho số
                                    is_diff = True
                            elif isinstance(v1_check, str):
                                # So sánh string có strip khoảng trắng thừa
                                if v1_check.strip() != v2_check.strip(): 
                                    is_diff = True
                            else:
                                # Các kiểu khác (datetime, boolean...)
                                if v1_check != v2_check:
                                    is_diff = True
                        
                        if is_diff:
                            v1_str = str(val_old) if val_old is not None else ""
                            v2_str = str(val_new) if val_new is not None else ""
                            
                            addr = ws_old.Cells(excel_r_old, c_idx+1).Address.replace("$", "")
                            report.append(CellDiff(idx_counter, "CELL", "MODIFIED", addr, v1_str, v2_str, f"Val changed"))
                            idx_counter += 1
                                
        return report, row_map, col_map

    def compare_shapes(self, shapes_old: Dict[int, ShapeData], shapes_new: Dict[int, ShapeData], 
                       row_map: Dict[int, int], col_map: Dict[int, int]) -> List[ShapeDiff]:
        
        report = []
        idx = 1
        
        def rc_to_addr_str(r, c):
            try:
                col_let = get_column_letter(c)
                return f"${col_let}${r}"
            except:
                return f"R{r}C{c}"

        # 1. Deleted
        for sid, s_old in shapes_old.items():
            if sid not in shapes_new:
                report.append(ShapeDiff(idx, sid, s_old.name, "DELETED", 0,0,0,0, s_old.anchor_address, "N/A", "N/A", 
                                        s_old.rel_left, s_old.rel_top, 0, 0))
                idx += 1
                
        # 2. New
        for sid, s_new in shapes_new.items():
            if sid not in shapes_old:
                report.append(ShapeDiff(idx, sid, s_new.name, "NEW", 0,0,0,0, "N/A", "N/A", s_new.anchor_address, 
                                        0, 0, s_new.rel_left, s_new.rel_top))
                idx += 1
                
        # 3. Common
        for sid, s_old in shapes_old.items():
            if sid in shapes_new:
                s_new = shapes_new[sid]
                
                exp_anchor_row = row_map.get(s_old.anchor_row, s_old.anchor_row)
                exp_anchor_col = col_map.get(s_old.anchor_col, s_old.anchor_col)
                
                exp_anchor_str = rc_to_addr_str(exp_anchor_row, exp_anchor_col)
                
                shift_x = s_new.rel_left - s_old.rel_left
                shift_y = s_new.rel_top - s_old.rel_top
                diff_w = s_new.width - s_old.width
                diff_h = s_new.height - s_old.height
                
                verdict = "MATCH"
                
                if (s_new.anchor_row != exp_anchor_row) or (s_new.anchor_col != exp_anchor_col):
                        verdict = "MOVED (Anchor Shift)"
                elif abs(shift_x) > self.tolerance or abs(shift_y) > self.tolerance:
                    verdict = "MOVED"
                elif abs(diff_w) > self.tolerance or abs(diff_h) > self.tolerance:
                    verdict = "RESIZED"

                report.append(ShapeDiff(idx, sid, s_old.name, verdict, shift_x, shift_y, diff_w, diff_h, 
                                        s_old.anchor_address, exp_anchor_str, s_new.anchor_address,
                                        s_old.rel_left, s_old.rel_top, s_new.rel_left, s_new.rel_top))
                idx += 1
                
        return report

# --- GUI CLASS ---

class AppUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title(f"{PROJECT_NAME} - 2.1") # [UPDATED] Version 2.1
        self.geometry("700x400") 
        self.minsize(700, 400)
        
        self.sheet_list_thread = None
        self.current_lang = "English" # Default Language
        
        self._setup_ui()
        self.update_ui_text() # Initial text set
        
    def _setup_ui(self):
        # Grid Configuration
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0) # Header
        self.grid_rowconfigure(1, weight=0) # Settings
        self.grid_rowconfigure(2, weight=1) # Spacer 
        self.grid_rowconfigure(3, weight=0) # Footer
        
        # --- A. HEADER SECTION ---
        self.frame_header = ctk.CTkFrame(self, corner_radius=10, border_width=1, border_color=CARD_BORDER_COLOR)
        self.frame_header.grid(row=0, column=0, padx=15, pady=15, sticky="ew")
        self.frame_header.grid_columnconfigure(1, weight=1)
        
        # Row 0: Old File
        self.lbl_old = ctk.CTkLabel(self.frame_header, text="Source (Old):", font=BOLD_FONT)
        self.lbl_old.grid(row=0, column=0, padx=10, pady=(15,5), sticky="w")
        
        self.btn_browse_old = ctk.CTkButton(self.frame_header, text="📂 Browse", width=100, font=MAIN_FONT, command=lambda: self.browse_file('old'))
        self.btn_browse_old.grid(row=0, column=0, padx=10, pady=(15,5), sticky="e") 
        
        self.entry_old = ctk.CTkEntry(self.frame_header, placeholder_text="Path...", font=MAIN_FONT)
        self.entry_old.grid(row=0, column=1, padx=5, pady=(15,5), sticky="ew")
        
        self.btn_open_old = ctk.CTkButton(self.frame_header, text="↗", width=30, font=MAIN_FONT, command=lambda: self.open_file_os(self.entry_old.get()))
        self.btn_open_old.grid(row=0, column=2, padx=5, pady=(15,5))
        
        self.cbo_sheet_old = ctk.CTkComboBox(self.frame_header, width=150, values=["Select File First"], font=MAIN_FONT)
        self.cbo_sheet_old.grid(row=0, column=3, padx=10, pady=(15,5), sticky="w")
        
        # Row 1: New File
        self.lbl_new = ctk.CTkLabel(self.frame_header, text="Target (New):", font=BOLD_FONT)
        self.lbl_new.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.btn_browse_new = ctk.CTkButton(self.frame_header, text="📂 Browse", width=100, font=MAIN_FONT, command=lambda: self.browse_file('new'))
        self.btn_browse_new.grid(row=1, column=0, padx=10, pady=5, sticky="e")

        self.entry_new = ctk.CTkEntry(self.frame_header, placeholder_text="Path...", font=MAIN_FONT)
        self.entry_new.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.btn_open_new = ctk.CTkButton(self.frame_header, text="↗", width=30, font=MAIN_FONT, command=lambda: self.open_file_os(self.entry_new.get()))
        self.btn_open_new.grid(row=1, column=2, padx=5, pady=5)
        self.cbo_sheet_new = ctk.CTkComboBox(self.frame_header, width=150, values=["Select File First"], font=MAIN_FONT)
        self.cbo_sheet_new.grid(row=1, column=3, padx=10, pady=5, sticky="w")
        
        # Row 2: Output
        self.lbl_out = ctk.CTkLabel(self.frame_header, text="Output Folder:", font=BOLD_FONT)
        self.lbl_out.grid(row=2, column=0, padx=10, pady=(5,15), sticky="w")
        self.btn_browse_out = ctk.CTkButton(self.frame_header, text="📂 Browse", width=100, font=MAIN_FONT, command=self.browse_folder)
        self.btn_browse_out.grid(row=2, column=0, padx=10, pady=(5,15), sticky="e")

        self.entry_out = ctk.CTkEntry(self.frame_header, placeholder_text="Default...", font=MAIN_FONT)
        self.entry_out.grid(row=2, column=1, padx=5, pady=(5,15), sticky="ew")
        self.btn_open_out = ctk.CTkButton(self.frame_header, text="↗", width=30, font=MAIN_FONT, command=lambda: self.open_file_os(self.entry_out.get()))
        self.btn_open_out.grid(row=2, column=2, padx=5, pady=(5,15))
        
        # --- B. CONFIGURATION ---
        self.frame_config = ctk.CTkFrame(self, corner_radius=10, border_width=1, border_color=CARD_BORDER_COLOR)
        self.frame_config.grid(row=1, column=0, padx=15, pady=5, sticky="ew")
        self.frame_config.grid_columnconfigure(0, weight=1)
        self.frame_config.grid_columnconfigure(1, weight=1)
        
        # Left: Scan Scope & Report Mode
        self.frame_scope = ctk.CTkFrame(self.frame_config, fg_color="transparent")
        self.frame_scope.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        
        # Row 0: Limit Scan
        self.chk_limit = ctk.CTkCheckBox(self.frame_scope, text="Limit Scan Range", font=MAIN_FONT, command=lambda: self.toggle_range_input(self.cbo_scope_type.get()))
        self.chk_limit.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        self.cbo_scope_type = ctk.CTkComboBox(self.frame_scope, 
                                                                values=["Whole Sheet", "Columns Only", "Rows Only", "Specific Range"], 
                                                                state="disabled", font=MAIN_FONT,
                                                                command=self.toggle_range_input)
        self.cbo_scope_type.grid(row=0, column=1, padx=5, pady=5)
        
        self.entry_range = ctk.CTkEntry(self.frame_scope, placeholder_text="e.g. A1:H50", state="disabled", font=MAIN_FONT)
        self.entry_range.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Row 1: Report Mode (Moved Here)
        self.lbl_report_mode = ctk.CTkLabel(self.frame_scope, text="Report Mode:", font=MAIN_FONT)
        self.lbl_report_mode.grid(row=1, column=0, padx=5, pady=(10,0), sticky="w")
        self.cbo_report_type = ctk.CTkComboBox(self.frame_scope, values=["Full Report", "Changes Only"], font=MAIN_FONT)
        self.cbo_report_type.grid(row=1, column=1, columnspan=2, padx=5, pady=(10,0), sticky="ew")
        
        # Right: Advanced & Flags
        self.frame_adv = ctk.CTkFrame(self.frame_config, fg_color="transparent")
        self.frame_adv.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        
        # Row 0: Force Kill (Moved to Right Col)
        self.chk_reset = ctk.CTkCheckBox(self.frame_adv, text="Force Kill Excel", font=MAIN_FONT)
        self.chk_reset.grid(row=0, column=0, padx=10, pady=5, sticky="w")

        # Row 1: Highlight (Moved to Right Col)
        self.chk_highlight = ctk.CTkCheckBox(self.frame_adv, text="Highlight Changes", font=MAIN_FONT)
        self.chk_highlight.grid(row=1, column=0, padx=10, pady=(10,0), sticky="w")
        self.chk_highlight.select() # [UPDATED] Mặc định tích chọn
        
        # NOTE: Tolerance UI is removed (Hidden), defaulted to constant in logic.
        
        # --- C. ACTION ---
        self.btn_start = ctk.CTkButton(self, text="START COMPARISON", 
                                                             fg_color=ACCENT_COLOR, hover_color=HOVER_COLOR,
                                                             height=55, font=("Segoe UI", 16, "bold"), corner_radius=8,
                                                             command=self.start_process)
        self.btn_start.grid(row=2, column=0, padx=15, pady=20, sticky="ew")
        
        # --- D. STATUS BAR (FOOTER) ---
        self.frame_footer = ctk.CTkFrame(self, fg_color=STATUS_BAR_BG, corner_radius=0, height=35)
        self.frame_footer.grid(row=3, column=0, padx=0, pady=0, sticky="ew")
        
        # Left Side: Status Text
        self.lbl_status = ctk.CTkLabel(self.frame_footer, text=" Ready.", font=("Segoe UI", 11, "bold"), 
                                                             text_color="#ECF0F1", anchor="w")
        self.lbl_status.pack(side="left", fill="x", expand=True, padx=10, pady=5)
        
        # Right Side: Version | Language | Help
        ctk.CTkLabel(self.frame_footer, text=f"v2.1 | {AUTHOR_ID} ", font=("Segoe UI", 10), # [UPDATED] v2.1
                             text_color="#95A5A6").pack(side="right", padx=10)
        
        self.cbo_lang = ctk.CTkComboBox(self.frame_footer, width=100, values=list(LANGUAGES.keys()), 
                                                              font=("Segoe UI", 11), command=self.change_language)
        self.cbo_lang.set("English")
        self.cbo_lang.pack(side="right", padx=5)
        
        self.btn_help = ctk.CTkButton(self.frame_footer, text="❓ Help", width=70, font=("Segoe UI", 11),
                                                            fg_color="#34495E", hover_color="#4E5D6C",
                                                            command=self.show_help_popup)
        self.btn_help.pack(side="right", padx=5)

    # --- UI EVENT HANDLERS ---
    
    def change_language(self, choice):
        self.current_lang = choice
        self.update_ui_text()

    def update_ui_text(self):
        text_data = LANGUAGES[self.current_lang]
        
        # Labels & Buttons
        self.lbl_old.configure(text=text_data["source"])
        self.lbl_new.configure(text=text_data["target"])
        self.btn_browse_old.configure(text=text_data["browse"])
        self.btn_browse_new.configure(text=text_data["browse"])
        self.btn_browse_out.configure(text=text_data["browse"])
        self.btn_open_old.configure(text=text_data["open"])
        self.btn_open_new.configure(text=text_data["open"])
        self.btn_open_out.configure(text=text_data["open"])
        
        self.entry_old.configure(placeholder_text=text_data["placeholder_old"])
        self.entry_new.configure(placeholder_text=text_data["placeholder_new"])
        
        self.lbl_out.configure(text=text_data["output"])
        self.entry_out.configure(placeholder_text=text_data["placeholder_out"])
        
        self.chk_limit.configure(text=text_data["limit_scan"])
        self.cbo_scope_type.configure(values=text_data["scope_types"])
        self.cbo_scope_type.set(text_data["scope_types"][0])
        
        # Moved Options
        self.chk_reset.configure(text=text_data["force_kill"])
        self.lbl_report_mode.configure(text=text_data["lbl_report_mode"])
        self.cbo_report_type.configure(values=text_data["report_modes"])
        self.cbo_report_type.set(text_data["report_modes"][0])
        self.chk_highlight.configure(text=text_data["chk_highlight"])
        
        self.btn_start.configure(text=text_data["start_btn"])
        self.lbl_status.configure(text=text_data["status_ready"])
        self.btn_help.configure(text=text_data["help_btn"])
        
        if self.cbo_sheet_old.get() in ["Select File First", "Chọn file trước", "ファイルを選択してください"]:
            self.cbo_sheet_old.set(text_data["sel_file_first"])
            self.cbo_sheet_old.configure(values=[text_data["sel_file_first"]])
            
        if self.cbo_sheet_new.get() in ["Select File First", "Chọn file trước", "ファイルを選択してください"]:
             self.cbo_sheet_new.set(text_data["sel_file_first"])
             self.cbo_sheet_new.configure(values=[text_data["sel_file_first"]])

    def show_help_popup(self):
        """Displays the help popup with multi-language support."""
        txt_data = LANGUAGES[self.current_lang]
        
        top = ctk.CTkToplevel(self)
        top.title(txt_data["help_title"])
        top.geometry("500x400")
        top.resizable(False, False)
        top.grab_set() 
        
        lbl_title = ctk.CTkLabel(top, text=txt_data["help_title"], font=("Segoe UI", 16, "bold"))
        lbl_title.pack(pady=10)
        
        textbox = ctk.CTkTextbox(top, font=("Segoe UI", 12), wrap="word")
        textbox.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        textbox.insert("1.0", txt_data["help_content"])
        textbox.configure(state="disabled")

    def toggle_range_input(self, choice=None):
        if choice is None: choice = self.cbo_scope_type.get()
        lang_data = LANGUAGES[self.current_lang]
        
        is_limit = self.chk_limit.get() == 1
        
        if is_limit:
            self.cbo_scope_type.configure(state="normal")
            
            idx = -1
            if choice in lang_data["scope_types"]:
                idx = lang_data["scope_types"].index(choice)
            
            if idx == 0: # Whole Sheet
                self.entry_range.configure(state="disabled")
            else:
                self.entry_range.configure(state="normal")
                if idx == 1: # Col
                    self.entry_range.configure(placeholder_text=lang_data["ph_range_col"])
                elif idx == 2: # Row
                    self.entry_range.configure(placeholder_text=lang_data["ph_range_row"])
                else: # Specific
                    self.entry_range.configure(placeholder_text=lang_data["ph_range_spec"])
        else:
            self.cbo_scope_type.configure(state="disabled")
            self.entry_range.configure(state="disabled")

    def browse_file(self, target):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xlsb")])
        if path:
            path = os.path.normpath(path)
            if target == 'old':
                self.entry_old.delete(0, "end")
                self.entry_old.insert(0, path)
                self.load_sheets_lazy(path, self.cbo_sheet_old)
            else:
                self.entry_new.delete(0, "end")
                self.entry_new.insert(0, path)
                self.load_sheets_lazy(path, self.cbo_sheet_new)

    def browse_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.entry_out.delete(0, "end")
            self.entry_out.insert(0, os.path.normpath(path))

    def open_file_os(self, path):
        if path and os.path.exists(path):
            os.startfile(path)

    def load_sheets_lazy(self, path, combo_widget):
        def _worker():
            try:
                wb = openpyxl.load_workbook(path, read_only=True, keep_links=False)
                sheets = wb.sheetnames
                wb.close()
                combo_widget.configure(values=sheets)
                combo_widget.set(sheets[0])
            except Exception as e:
                print(f"Lazy load error: {e}")
                
        threading.Thread(target=_worker, daemon=True).start()

    def log(self, msg):
        print(f"[LOG] {msg}")
        self.lbl_status.configure(text=f"{msg}")

    # --- MAIN PROCESS THREAD ---

    def start_process(self):
        txt = LANGUAGES[self.current_lang]["processing"]
        self.btn_start.configure(state="disabled", text=txt, fg_color="#7F8C8D")
        threading.Thread(target=self.run_logic, daemon=True).start()

    def run_logic(self):
        engine = ExcelEngine()
        txt_data = LANGUAGES[self.current_lang]
        
        try:
            # 1. Inputs
            p_old = self.entry_old.get()
            p_new = self.entry_new.get()
            # Mặc định ra Downloads
            p_out = self.entry_out.get() or os.path.join(os.path.expanduser("~"), "Downloads")
            
            sh_old_name = self.cbo_sheet_old.get()
            sh_new_name = self.cbo_sheet_new.get()
            
            # Default Tolerance (Hidden from UI now)
            tol = TOLERANCE_DEFAULT
                
            # Get Report Logic
            current_mode_str = self.cbo_report_type.get()
            is_only_diffs = (current_mode_str == txt_data["report_modes"][1]) # Index 1 is Changes Only
            is_highlight = (self.chk_highlight.get() == 1)

            # 2. Init
            self.log("Opening Excel securely...")
            engine.robust_open(p_old, p_new, force_reset=(self.chk_reset.get() == 1))
            
            ws_old = engine.get_sheet_by_name(engine.wb_old, sh_old_name)
            ws_new = engine.get_sheet_by_name(engine.wb_new, sh_new_name)
            
            # 3. Determine Range
            scan_range = None
            if self.chk_limit.get() == 1:
                scan_type = self.cbo_scope_type.get()
                if scan_type != txt_data["scope_types"][0]: 
                    scan_range = self.entry_range.get() 
            
            # 4. Compare Grids & Cells
            comp = Comparator(tol)
            cell_report, r_map, c_map = comp.compare_grids_and_cells(ws_old, ws_new, engine, log_func=self.log)
            
            # 5. Extract & Compare Shapes
            self.log(f"Scanning Old Shapes...")
            shapes_old = engine.extract_shapes(ws_old, "Source (Old)", scan_range, log_func=self.log)
            
            self.log(f"Scanning New Shapes...")
            shapes_new = engine.extract_shapes(ws_new, "Target (New)", scan_range, log_func=self.log)
            
            self.log(f"Comparing Shapes...")
            shape_report = comp.compare_shapes(shapes_old, shapes_new, r_map, c_map)
            
            # 6. Generate Report
            self.log("Generating Excel Report with Copies...")
            out_file = engine.create_report_workbook(p_out, cell_report, shape_report, 
                                                     ws_old, ws_new, 
                                                     only_diffs=is_only_diffs, 
                                                     highlight_changes=is_highlight)
            
            self.log(txt_data["status_done"])
            
            # =========================================================================
            # [CẬP NHẬT] HỘP THOẠI YES/NO ĐỂ MỞ FILE
            # =========================================================================
            
            # Chuẩn bị nội dung thông báo theo ngôn ngữ
            if self.current_lang == "Tiếng Việt":
                msg_title = "So sánh hoàn tất"
                msg_body = f"{txt_data['msg_success']}\n{out_file}\n\nBạn có muốn mở file báo cáo ngay không?"
            elif self.current_lang == "日本語":
                msg_title = "比較完了"
                msg_body = f"{txt_data['msg_success']}\n{out_file}\n\n今すぐレポートを開きますか？"
            else: # English (Default)
                msg_title = "Comparison Complete"
                msg_body = f"{txt_data['msg_success']}\n{out_file}\n\nDo you want to open the report now?"

            # Hiển thị hộp thoại Yes/No
            should_open = messagebox.askyesno(msg_title, msg_body)
            
            if should_open:
                try:
                    os.startfile(out_file)
                except Exception as open_err:
                    messagebox.showerror("Error", f"Could not open file:\n{str(open_err)}")

        except Exception as e:
            import traceback
            traceback.print_exc() 
            messagebox.showerror("Error", f"{txt_data['msg_error']}{str(e)}")
            self.log(txt_data["status_error"])
        finally:
            engine.cleanup()
            self.btn_start.configure(state="normal", text=txt_data["start_btn"], fg_color=ACCENT_COLOR)

if __name__ == "__main__":
    app = AppUI()
    app.mainloop()
