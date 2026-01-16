import xlsxwriter
import openpyxl
import zipfile
import os

filename = 'test_shapes.xlsx'

def create_excel_with_shapes():
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    
    # Add a textbox (which is a shape)
    worksheet.insert_textbox('B2', 'This is a textbox', {'width': 200, 'height': 100})
    
    # Add another one
    worksheet.insert_textbox('E5', 'Another shape', {'width': 100, 'height': 50})
    
    # Add some text to ensure content is also there
    worksheet.write('A1', 'Hello')
    worksheet.write('C3', 'World')
    
    workbook.close()
    print(f"Created {filename}")

def read_with_openpyxl():
    print("\n--- OpenPyxl Analysis ---")
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        # Check for images/shapes
        # Note: openpyxl support for shapes is limited, often grouped with images
        if hasattr(ws, '_images'):
            print(f"Found {len(ws._images)} images/shapes via _images")
            for img in ws._images:
                print(f"Image: {img.anchor}")
        
        if hasattr(ws, '_drawing'):
             print(f"Drawing object: {ws._drawing}")
             
    except Exception as e:
        print(f"OpenPyxl error: {e}")

def read_xml_structure():
    print("\n--- XML Analysis ---")
    try:
        with zipfile.ZipFile(filename, 'r') as z:
            # List all files to find drawings
            stats = z.namelist()
            drawing_files = [f for f in stats if 'drawing' in f]
            print(f"Drawing files found: {drawing_files}")
            
            for df in drawing_files:
                content = z.read(df)
                print(f"Content of {df} (first 500 chars):")
                print(content[:500])
                
    except Exception as e:
        print(f"Zip error: {e}")


if __name__ == "__main__":
    print(f"XlsxWriter version: {xlsxwriter.__version__}")
    print(f"OpenPyxl version: {openpyxl.__version__}")
    try:
        create_excel_with_shapes()
    except Exception as e:
        print(f"Creation failed: {e}")
        # Identify attributes
        try:
             import xlsxwriter
             wb = xlsxwriter.Workbook('dummy.xlsx')
             ws = wb.add_worksheet()
             print(f"Worksheet attributes: {[a for a in dir(ws) if 'insert' in a]}")
             wb.close()
        except:
            pass

    if os.path.exists(filename):
        read_with_openpyxl()
        read_xml_structure()

