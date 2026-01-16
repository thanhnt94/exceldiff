import openpyxl
from openpyxl.styles import PatternFill, Font, Color
from core.data_types import DiffResult, DiffType

class ExcelReportGenerator:
    def __init__(self, file_b_path: str, output_path: str):
        self.file_b_path = file_b_path
        self.output_path = output_path
        
    def generate(self, result: DiffResult):
        wb = openpyxl.load_workbook(self.file_b_path)
        ws = wb.active # Assume active sheet
        
        # Styles
        fill_changed = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
        fill_inserted = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Green
        fill_moved = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # Orange
        
        deleted_items = []
        
        for item in result.items:
            if item.item_type == "Cell":
                if item.diff_type == DiffType.CHANGED:
                    # Location is like "A1 -> A1" or just "A1".
                    # Comparator output: "A3 -> A4". We want the target coord in B.
                    # My comparator output format for CHANGED was "A3 -> A4".
                    # Let's parse it or fix comparator to provide clean target coord.
                    # Actually, let's fix comparator in next step to be cleaner, 
                    # but for now, parse it.
                    loc = item.location.split("->")[-1].strip()
                    try:
                        cell = ws[loc]
                        cell.fill = fill_changed
                        # Add comment? openpyxl comments are a bit complex to position, 
                        # but simple string comment is okay.
                    except:
                        pass
                elif item.diff_type == DiffType.INSERTED:
                    loc = item.location
                    try:
                        cell = ws[loc]
                        cell.fill = fill_inserted
                        
                        # Highlighting entire row?
                        # If detail says "Row inserted", we should highlight the whole row span or at least the used cells.
                        if "Row inserted" in item.details:
                            # Apply to reasonable range, e.g. A:End
                            # Use cell.row
                            for col_idx in range(1, 20): # Hardcoded max col for visual
                                try:
                                    ws.cell(row=cell.row, column=col_idx).fill = fill_inserted
                                except: pass
                    except:
                        pass
                elif item.diff_type == DiffType.DELETED:
                    deleted_items.append(item)

            elif item.item_type == "Shape":
                # Highlighting shapes is hard via openpyxl.
                # We will list them in the "Summary" sheet.
                pass

        # Create Summary Sheet
        ws_summary = wb.create_sheet("Diff Summary", 0)
        ws_summary.append(["Type", "Location", "Details", "Old Value", "New Value"])
        
        for item in result.items:
            ws_summary.append([
                item.diff_type.value,
                item.location,
                item.details,
                str(item.old_value) if item.old_value else "",
                str(item.new_value) if item.new_value else ""
            ])
            
        wb.save(self.output_path)
